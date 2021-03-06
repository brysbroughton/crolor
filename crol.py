from bs4 import BeautifulSoup as bs
from urlparse import urlparse
from datetime import datetime
import re, os, sys, httplib, urllib
import smtplib
import email,email.encoders,email.mime.base
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting import ColorScaleRule, CellIsRule, FormulaRule



class GenericType(object):
    """
    General template for associating properties to actions.
    To be extended by other crol classes.
    
    type: generic
    """

    def setprop(self, key, val):
        if self.props.has_key(key):
            self.props[key] = val
            setattr(self, key, val)
        else:
            raise Exception('%s has no property: %s' % (self.props['type'], key))

    def getprop(self, key):
        if self.props.has_key(key):
            return self.props[key]
        else:
            raise Exception('%s has no property: %s' % (self.props['type'], key))

    def listprops(self):
        for key, val in self.props.iteritems():
            print "%s : %s" % (key, val)
    
    def __extend_props__(self, key, val=None):
        self.props[key] = val
        setattr(self, key, val)
    
    def __init__(self, **kwargs):
        self.props = self.props or {'type' : 'generic'}
        if not self.props.has_key('type'):
            raise Exception('Generic subtype must have a type specified in properties')
        for key, val in self.props.iteritems():
            self.setprop(key, val)
        self.args = kwargs
        self.__use_args__()
    
    def __use_args__(self):
        for key, val in self.args.iteritems():
            if self.props.has_key(key):
                self.setprop(key, val)
            else:
                raise Exception('%s object has no property: %s' % (self.props['type'], key))


class Registration(GenericType):
    """
    Object that associates a department to a website and actions.
    
    department: department object
    site: dep.otc.edu/artment
    actions: list of action objects
    """
    
    def __init__(self, kwargs={}):
        self.props = {
            'type' : 'registration',
            'department' : None,
            'site' : None,
            'actions' : []
        }
        
        super(Registration, self).__init__(**kwargs)
        
        if self.props['department']:
            setattr(self, 'department', Department(self.props['department']))


class Department(GenericType):
    """
    Object handles information for an orgnaizational entity associated with a crawl
    type : department
    name : department_name
    main_email : department@otc.edu
    email_group : [an@email.com, another@email.edu]
    """
    
    def __init__(self, kwargs={}):
        self.props = {
            'type' : 'department',
            'name' : None,
            'main_email' : None,
            'email_group' : []
        }
        
        super(Department, self).__init__(**kwargs)


class Registry(GenericType):
    """
    Object handles the listing of sites to crawl and associates them to departments.
    """
    
    def __init__(self, kwargs={}):
        self.props = {
            'type' : 'registry',
            'registrations' : []
        }
        
        super(Registry, self).__init__(**kwargs)
        
        if self.props['registrations']:
            setattr(self, 'registrations', [Registration(r) for r in self.props['registrations']])

class CrawlReport(GenericType):
    """
    Object instatiated 1-1 with a crawl job.
    """
    
    def __init__(self, kwargs={}):
        self.props = {
            'type' : 'crawl_report',
            'seed_url' : None,
            'page_reports' : [],
            'url_reports' : []#not sure if use
        }

        super(CrawlReport, self).__init__(**kwargs)
        
        if self.props['page_reports']:
            setattr(self, 'page_reports', [PageReport(r) for r in self.props['page_reports']])
        
        if self.props['url_reports']:
            setattr(self, 'url_reports', [UrlReport(r) for r in self.props['url_reports']])
        
    def addreport(self, report):
        if report.type == 'page_report':
            self.page_reports.add(report)
        elif report.type == 'url_report':
            self.url_reports.append(report)
            self.html_chunk += self.log.buildrow([report.status, report.reason, report.mimetype, report.url, report.parent_url])


class PageReport(GenericType):
    """
    Object for handling the reports of all links on its page
    """
    
    def __init__(self, kwargs={}):
        self.props = {
            'type' : 'page_report',
            'page_url' : None,
            'page_status' : None,
            'url_reports' : []
        }
        
        super(PageReport, self).__init__(**kwargs)
        
        if self.props['url_reports']:
            setattr(self, 'url_reports', [UrlReport(r) for r in self.props['url_reports']])


class UrlReport(GenericType):
    """
    Object for handling url and the response received when it is requested.
    """
    
    def __init__(self, kwargs={}):
        self.props = {
            'type' : 'url_report',
            'url' : None,
            'linked_from' : None
        }
        
        super(UrlReport, self).__init__(**kwargs)
        


class Node(GenericType):
    """
    Handles a url request information.
    """
    
    def __init__(self, kwargs={}):
        self.props = {
            'type' : 'node',
            'url' : None,
            'response' : None,#replace by response object from httplib
            'urlparse' : None,
            'headers' : None,
            'mimetype' : None,
            'status' : None,
            'reason' : None,
            'links' : [],
            'parent': None,#'HEAD'
            'children' : []
        }
        
        super(Node, self).__init__(**kwargs)
        
        #if not self.url:
        #    raise Exception("Node object must be created with a url")
        print 'link:', self.url #to show progess in console
        
        self.setprop('url', self.normalize(self.getprop('url')))#first step is to normalize all urls
        print self.url #to show progess in console
        self.setprop('urlparse', urlparse(self.getprop('url')))
        if self.urlparse.scheme in ['http', 'https']:
            self.request()
        elif self.urlparse.scheme == 'mailto':
            self.checkemail()
    
    def normalize(self, link):
        """
        Take a string link similar to /admissions
        Return a valid url like http://otc.edu/admissions
        Use urlparse to get the pieces of the link.
        If essential components (scheme, netloc) are missing, attempt to use those from parent node
        """
        
        if link is None or len(link) == 0:
            return ''
            
        link = link.strip()#remove whitespace from ends
        
        new_parsed = urlparse(link)

        if new_parsed.scheme == 'mailto':
            return link

        #relative paths are tricky
        new_path = new_parsed.path
        new_link = []
        
        try:
            if new_path.startswith('/'):
                pass
            else:#link with no leading slash should be sub-link of current directory
                if self.urlparse:
                    old_path_bits = re.split('/', self.urlparse.path)
                    if '.' in old_path_bits[-1]:#url ends with a filename
                        new_path_bits = old_path_bits[:-1] + [new_path]
                        new_path = '/'.join(new_path_bits)
                    else:
                        new_path = (self.urlparse.path + '/' + new_parsed.path).replace('//','/')
            new_link = [
                (new_parsed.scheme or self.urlparse.scheme) + '://',
                new_parsed.netloc or self.urlparse.netloc,
                new_path,
                #';'+new_parsed.params,
                '?'+new_parsed.query if new_parsed.query else '',
                #'#'+new_parsed.fragment
            ]
        except AttributeError as AEX:
            raise Exception("Could not normalize url. Url is mal-formed, or a relative url without a parent node. ORIGINAL ERROR: "+AEX.message)
        
        return ''.join(new_link)
    
    def request(self):
        """
        Request the url provided to the constructor.
        Set node url status code and reason.
        Call scape() and set node links list.
        """
        
        #connect and collect node.url info
        if self.urlparse.scheme == 'http':
            conn = httplib.HTTPConnection(self.urlparse.netloc)
        elif self.urlparse.scheme == 'https':
            conn = httplib.HTTPSConnection(self.urlparse.netloc)
        else:
            raise Exception("Node attempted to request unsupported protocol: "+self.url)
                    
        conn.request('GET',self.urlparse.path)
        response = conn.getresponse()
        self.setprop('response', response)
        self.setprop('status', response.status)
        self.setprop('reason', response.reason)
        self.setprop('headers', response.getheaders())
        #check for text/html mime type to scrape html
        self.setprop('mimetype', response.getheader('content-type'))
        if self.mimetype is not None and 'text/html' in self.mimetype:
            if str(self.status) != '404':
                html_response = response.read()
                self.setprop('links', self.scrape(html_response))
    
    def scrape(self, html):
        """
        Use beautiful soup to get all the links off of the page.
        Return scraped links in set form.
        
        If the header includes a redirect, the location will be added to the nodes links,
        the same as a link in the response body. A correctly configured server will return
        the same link in the body as in the header redirect.
        """
        
        links = []
        
        #check headers for redirects
        redirects = filter(lambda x: x[0] == 'location', self.headers)
        links.extend([x[1] for x in redirects])
        
        #scrape response body
        soup = bs(html)
        metalinks = soup.findAll('meta', attrs={'http-equiv':True})
        for m in metalinks:
            index = str(m).find('url=')
            end = str(m).find('"',index, len(str(m)))
            if index != -1:
                link = str(m)[index+4:end]
                links.append(link)
        
        attrs = ['background', 'cite', 'codebase', 'href', 'longdesc', 'src']
            
        for a in attrs:
            links.extend(
                map(lambda x: x[a], soup.findAll(**{a:True}))# **{} unzips dictionary to a=True
            )
        
        return links
        
    def checkemail(self):
        """
        Called on node that contains an email link.
        Evaluates the link for correctness and sets internal properties accordingly
        """
        valid = re.match('^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[A-Za-z]{2,4}$', self.urlparse.path)
        self.setprop('status', '200' if valid else '404')
        self.setprop('reason', 'Address correctly formatted' if valid else 'invalid email address')


class Crawl(GenericType):
    """
    Executes site crawl by creating and maintaining Node tree.
    """
    
    def __init__(self, kwargs={}):
        self.props = {
            'type' : 'crawler',
            'seed_url' : None,
            'node_tree' : None,#'HEAD'
            'visited_urls' : set([]),
            'log' : None,
        }
        
        super(Crawl, self).__init__(**kwargs)
    
    def start(self, funcin=None):
        """
        Begin the crawl process from url seed
        """
        head = Node({'url':self.seed_url, 'parent':'HEAD'})
        self.setprop('node_tree', head)
        if funcin: funcin(head)
        self.reccrawl(head, funcin)
    
    def reccrawl(self, node, funcin=None):
        self.visited_urls.add(node.url)
        
        for l in node.links:
            new_url = None
            try:
                new_url = node.normalize(l)
            except IOError as error:
                print 'Could not normalize url: ', l#error
                
            if new_url:
                new_node = Node({'url':new_url})
            else:
                new_node = Node({'url':'', 'status':'404', 'reason':'Empty link'})
            new_node.setprop('parent', node)
            node.children.append(new_node)
            
            if funcin: funcin(new_node)
            
            if new_url not in self.visited_urls:
                if self.shouldfollow(new_url):
                    self.reccrawl(new_node, funcin)
    
    def shouldfollow(self, url):
        """
        Take node object, return boolean
        #don't crawl the same url 2x
        #only crawl urls within a subsite of the input seed
        """
        if url not in self.getprop('visited_urls'):
            url = urlparse(url)
            seed = urlparse(self.getprop('seed_url'))
            if url.netloc == seed.netloc and url.path[:len(seed.path)] == seed.path:
                return True
            else:
                return False
        else:
            return False

            
class Log(GenericType):
    """
    Generic text log handling for crawl job
    """
    
    def __init__(self, kwargs={}):
        self.props = {
            'type' : 'log',
            'path' : './',
            'filename' : 'croler',
            'endfilename' : '.log.txt',
            'head_text' : 'header\n',
            'foot_text' : 'footer',
            'filePointer' : None,
            "row_before" : '',
            "row_after" : '',
            "col_before" : '',
            "col_after" : '',
            "row_trim" : "right"
        }
        super(Log, self).__init__(**kwargs)
    
    def openfile(self):
        f = open(self.path+self.filename+self.endfilename, 'w')
        self.setprop('filePointer', f)
        self.writefile(self.head_text)
    
    def writefile(self, new_val):
        self.filePointer.write(new_val)
        self.filePointer.flush()
        os.fsync(self.filePointer)
    
    def closefile(self):
        self.writefile(self.foot_text)
        self.filePointer.close
    
    def writerow(self, row):
        string_bits = []
        string_bits.append(self.row_before)
        for col in row:
            string_bits.append(self.col_before)
            string_bits.append(col)
            string_bits.append(self.col_after)
        if self.row_trim:
            if self.row_trim == 'left':
                string_bits = string_bits[1:len(string_bits)]
            elif self.row_trim == 'right':
                string_bits = string_bits[0:len(string_bits)-1]
            elif self.row_trim == 'both':
                string_bits = string_bits[1:len(string_bits)-1]
            elif self.row_trim == 'none':
                pass
            else:
                raise Exception('The only values accepted for row_trim are: "left, right, both, or none". You entered %s' % self.row_trim )
        else:
            raise Exception('Property "row_trim" must have a value')
        string_bits.append(self.row_after)
        string_bits = map(str, string_bits)
        self.writefile(''.join(string_bits))
    
    
    

class WebLog(Log):
    """
    HTML log handler for crawl job
    """
    
    def __init__(self, kwargs={}):
        self.props = {
            'type' : 'weblog',
            'path' : './',
            'filename' : 'webLog',
            'endfilename' : '.log.html',
            'head_text' : '<!DOCTYPE html><html><head></head><body><table>',
            'foot_text' : '\n</table></body></html>',
            'filePointer' : None,
            'html_before' : '<!DOCTYPE html><html>\n<head></head>\n<body>',
            'html_after' : '\n</body></html>',
            'table_wrapper' : '\n<div class="table">',
            'row_wrapper' : '\n\n<div class="row">',
            'col_wrapper' : '\n<div class="col">',
            'wrapper_after' : '</div>',
            'html_chunks' : [],
            'default_headings' : []
        }
        GenericType.__init__(self, **kwargs)
        self.injectcss()
    
    def openfile(self):
        f = open(self.path+self.filename+self.endfilename, 'w')
        self.setprop('filePointer', f)
        self.writefile(self.html_before)
    
    def closefile(self):
        for h in self.html_chunks:
            self.writefile(h)
        self.writefile(self.html_after)
        self.filePointer.close
    
    def wrapchunk(self, html_chunk):
        wrapped_chunk = self.table_wrapper + html_chunk + self.wrapper_after
        self.html_chunks.append(wrapped_chunk)
        return wrapped_chunk
    
    def buildrow(self, row):
        """
        Builds a string of HTML code from the given row.
        Returns the HTML code as a string.
        """
        
        string_bits = []
        string_bits.append(self.row_before)
        for col in row:
            string_bits.append(self.col_before)
            string_bits.append(str(col).replace('<','&lt;').replace('>','&gt;'))
            string_bits.append(self.col_after)
        string_bits.append(self.row_after)
        string_bits = map(str, string_bits)
        self.writefile(''.join(string_bits))


class CsvLog(Log):
    """
    CSV log handler for crawl job
    """
    
    def __init__(self, kwargs={}):
        self.props = {
            'type' : 'log',
            'path' : './',
            'filename' : 'csvLog',
            'endfilename' : '.log.csv',
            'head_text' : 'header\n',
            'foot_text' : 'footer',
            'filePointer' : None,
            "row_before" : '',
            "row_after" : '\n',
            "col_before" : '',
            "col_after" : ',',
            "heading_row" : [],
            "row_trim" : "right"
        }
        GenericType.__init__(self, **kwargs)

    def headingrow(self, headings=None):
        if headings:
            self.writefile(self.row_before)
            for header in  headings:
                self.writefile(self.col_before)
                self.writefile(header)
                self.writefile(self.col_after)
        else:
            self.writefile(self.heading_row)
        self.writefile(self.row_after)

class ExcelLog(Log):
    def __init__(self, kwargs={}):
        self.props = {
            'type' : 'excelLog',
            'path' : './',
            'filename' : 'excelLog',
            'endfilename' : '.xlsx',
            'head_text' : 'header\n',
            'foot_text' : 'footer',
            'filePointer' : None,
            "row_before" : '',
            "row_after" : '',
            "col_before" : '',
            "col_after" : '',
            "heading_row" : [],
            "row_trim" : "none",
            "workbook" : None,
            "worksheet" : None,
            "numrows" : 1
        }
        GenericType.__init__(self, **kwargs)
    
    def openfile(self):
        wb = Workbook()
        ws = wb.active
        self.workbook = wb
        self.worksheet = ws
        self.headingrow(['STATUS', 'REASON', 'MIMETYPE', 'URL', 'PARENT'])

        
    def writefile(self):
        wb = self.workbook
        ws = self.worksheet
    
    def closefile(self):
        self.workbook.save(self.path+self.filename + self.endfilename)
    
    def writerow(self, vals):
        ws = self.workbook.active
        this_row = self.numrows
        this_col = 1
        #code_color = None
        code_color = PatternFill(start_color='00FFCC00', fill_type='solid')
        print "CODE: "
        print str(vals)[1]
        index = 1
        if str(vals)[1] == "'":
            index = 2
        if str(vals)[index] == '1':
            code_color = PatternFill(start_color='004472B9', end_color='004472B9', fill_type='solid') #blue
        elif str(vals)[index] == '2':
            code_color = PatternFill(start_color='004CA454', end_color='004CA454', fill_type='solid') #green
        elif str(vals)[index] == '3':
            code_color = PatternFill(start_color='00D49b00', end_color='00D49b00', fill_type='solid') #orange
        elif str(vals)[index] == '4':
            code_color = PatternFill(start_color='00BE4C39', end_color='00BE4C39', fill_type='solid') #red
        elif str(vals)[index] == '5':
            code_color = PatternFill(start_color='009351A6', end_color='009351A6', fill_type='solid') #purple
            
        for v in vals:
            cell = ws.cell(row = this_row, column = this_col)
            cell.value = v
            cell.fill = code_color
            if ws.column_dimensions[get_column_letter(this_col)].width < len(str(v)):
                ws.column_dimensions[get_column_letter(this_col)].width = len(str(v)) + 4
            this_col += 1
        self.numrows += 1
        self.worksheet = ws
        
    def headingrow(self, headings=None):
        ws = self.workbook.active
        if headings:
            this_col = 1
            this_row = self.numrows
            for header in  headings:
                cell = ws.cell(row = this_row, column = this_col)
                cell.value = header
                if ws.column_dimensions[get_column_letter(this_col)].width < len(str(header)):
                    ws.column_dimensions[get_column_letter(this_col)].width = len(str(header)) + 4
                this_col += 1
        else:
            this_col = 1
            this_row = self.numrows
            for header in  self.heading_row:
                cell = ws.cell(row = this_row, column = this_col)
                cell.value = header
                print "ELSE: "
                print cell.value
                if ws.column_dimensions[get_column_letter(this_col)].width < len(str(header)):
                    ws.column_dimensions[get_column_letter(this_col)].width = len(str(header)) + 4
                this_col += 1
        self.numrows += 1
        self.worksheet = ws


class Email(GenericType):
    def __init__(self, kwargs={}):
        self.props = {
        'type' : 'email',
        'msg_body' : '',
        'subject' : '',
        'from_address' : "",
        'to_address' : "",
        'cc_address' : '',
        'files' : [],
        'filename' : '',
        'smtp_server' : 'smtp.otc.edu',
        'mime_type' : "html"
        }
        super(Email, self).__init__(**kwargs)
    def send(self):
        has_cc = False
        addresses = [self.to_address]
        if self.props['to_address'] == "":
            raise Exception('to_address must be set in class: Email')
        elif self.props['from_address'] == "":
            raise Exception('from_address must be set in class: Email')
        else:
            #composing message using MIMEText
            msg = MIMEMultipart()
            msg['Subject'] = self.subject
            msg['From'] = self.from_address
            msg['To'] = self.to_address
            if self.cc_address != "":
                msg['CC'] = self.cc_address
                has_cc = True
                addresses.append(self.cc_address)
            if self.props['mime_type'] == "plain":
                message = MIMEText(self.msg_body, 'plain')
                msg.attach(message)
            else:
                message = MIMEText(self.msg_body, 'html')
                msg.attach(message)

            for f in self.files:
                fileMsg = email.mime.base.MIMEBase('application','octet-stream')
                fileMsg.set_payload(open(f, 'rb').read())
                email.encoders.encode_base64(fileMsg)
                fileMsg.add_header('Content-Disposition','attachment;filename=%s' % self.filename)
                msg.attach(fileMsg)
            
            
            
            #Email transmission with smtplib and OTC servers
            s = smtplib.SMTP(self.smtp_server)
            s.sendmail(self.from_address, addresses, msg.as_string())
            s.quit()
